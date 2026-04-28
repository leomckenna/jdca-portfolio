#!/usr/bin/env python3
"""
JDCA Portfolio — Yahoo Finance Data Collector
==============================================
Pulls data for every portfolio holding and saves it to an Excel file.

HOW TO RUN
----------
  1. Open a terminal in this folder
  2. Run:  python collect_data.py

HOW TO CHANGE THE DATE RANGE
-----------------------------
  Find the "DATE RANGE" section below and edit START_DATE and END_DATE.
  The format is always:  datetime(YEAR, MONTH, DAY)

WHAT YOU GET
------------
  An Excel file with 7 sheets:
    Info & Ratios       - snapshot of ~180 fields Yahoo Finance provides
                          (price, P/E, market cap, margins, beta, etc.)
    Daily Prices        - every trading day in the period with OHLCV,
                          adjusted close, dividends, and split columns
    Dividends           - only the days a dividend was paid
    Stock Splits        - only the days a split occurred
    Income Stmt (Q)     - last 4 quarters of income statements
    Balance Sheet (Q)   - last 4 quarters of balance sheets
    Cash Flow (Q)       - last 4 quarters of cash flows

  Note: quarterly financials always pull the most recent filings available,
  regardless of the date range you set above.

REQUIREMENTS
------------
  pip install yfinance pandas openpyxl
"""

# ---------------------------------------------------------------------------
# IMPORTS — you don't need to change anything here
# ---------------------------------------------------------------------------
import argparse
import warnings
from datetime import datetime, timedelta

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ===========================================================================
# DATE RANGE
# Change START_DATE and END_DATE to pull a different period.
# Format: datetime(YEAR, MONTH, DAY)
# ===========================================================================
START_DATE = datetime(2026, 1, 1)    # ← change this: start of the period
END_DATE   = datetime(2026, 3, 31)   # ← change this: end of the period


# ===========================================================================
# TICKERS
# Add or remove tickers from this list to change which companies are pulled.
# Use Yahoo Finance ticker symbols (e.g. "AAPL", "MSFT").
# Non-US tickers need their exchange suffix (e.g. "FLUI.ST" for Stockholm).
# ===========================================================================
TICKERS = [
    "LLY",       # Eli Lilly
    "NVO",       # Novo Nordisk
    "SNY",       # Sanofi
    "VRTX",      # Vertex Pharma
    "OTSKY",     # Otsuka Holdings
    "CRSP",      # CRISPR Therapeutics
    "SANA",      # Sana Biotech
    "EVO",       # Evotec SE
    "LCTX",      # Lineage Cell
    "HUMA",      # Humacyte
    "GNPX",      # Genprex
    "IPSC",      # Century Therapeutics
    "SABS",      # SAB Biotherapeutics
    "SEOVF",     # Sernova
    "NCEL",      # NewcelX
    "FLUI.ST",   # Fluicell (Stockholm)
    "NXTCL.ST",  # NextCell Pharma (Stockholm)
    "IMCR",      # Immunocore
    "CELZ",      # Creative Medical
    "ELDN",      # Eledon Pharma
    "ADOC.PA",   # Adocia (Paris)
    "XLV",       # Health Care Select Sector ETF (benchmark)
]


# ===========================================================================
# OUTPUT FILENAME
# The file will be saved in the same folder as this script.
# By default the name includes the date range so it's easy to identify.
# Change the string below if you want a fixed name, e.g. "Q1_2026_data.xlsx"
# ===========================================================================
OUTPUT_FILE = (
    f"jdca_data_{START_DATE.strftime('%Y%m%d')}_to_{END_DATE.strftime('%Y%m%d')}.xlsx"
)
# Uncomment the line below to use a fixed name instead:
# OUTPUT_FILE = "Q1_2026_data.xlsx"


# ---------------------------------------------------------------------------
# HELPER FUNCTIONS
# These handle small tasks used throughout the script.
# You don't need to change anything in this section.
# ---------------------------------------------------------------------------

def remove_timezone(index):
    """
    Yahoo Finance returns dates with timezone info attached.
    Excel can't handle that, so this strips the timezone off and zeroes the time component.
    """
    if isinstance(index, pd.DatetimeIndex):
        if index.tz is not None:
            index = index.tz_convert(None)
        return index.normalize()
    return index


def remove_timezone_from_df(df):
    """Strip timezone from every date column in a DataFrame."""
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                if df[col].dt.tz is not None:
                    df[col] = df[col].dt.tz_convert(None)
            except Exception:
                pass
    if isinstance(df.index, pd.DatetimeIndex) and df.index.tz is not None:
        df.index = df.index.tz_convert(None)
    return df


def build_info_row(ticker, info_dict):
    """
    Turn a ticker's .info dictionary into a single flat row.
    Puts the most useful fields first, then adds everything else.
    """
    row = {"Ticker": ticker}

    # These fields will appear first in the Info sheet (most commonly used)
    priority_fields = [
        "shortName", "longName", "sector", "industry", "country", "currency",
        "exchange", "quoteType", "marketCap", "enterpriseValue",
        "currentPrice", "previousClose", "open", "dayLow", "dayHigh",
        "fiftyTwoWeekLow", "fiftyTwoWeekHigh", "fiftyDayAverage", "twoHundredDayAverage",
        "volume", "averageVolume", "averageVolume10days",
        "trailingPE", "forwardPE", "priceToBook", "priceToSalesTrailing12Months",
        "trailingEps", "forwardEps", "bookValue",
        "dividendRate", "dividendYield", "exDividendDate", "payoutRatio",
        "fiveYearAvgDividendYield", "trailingAnnualDividendRate",
        "lastSplitFactor", "lastSplitDate",
        "beta", "shortRatio", "shortPercentOfFloat",
        "heldPercentInsiders", "heldPercentInstitutions",
        "sharesOutstanding", "sharesFloat",
        "totalRevenue", "revenuePerShare", "revenueGrowth",
        "grossProfits", "grossMargins", "operatingMargins", "profitMargins",
        "ebitdaMargins", "ebitda", "operatingCashflow", "freeCashflow",
        "totalCash", "totalCashPerShare", "totalDebt", "debtToEquity",
        "returnOnAssets", "returnOnEquity",
        "earningsGrowth", "earningsQuarterlyGrowth",
        "currentRatio", "quickRatio",
        "targetHighPrice", "targetLowPrice", "targetMeanPrice", "targetMedianPrice",
        "recommendationKey", "numberOfAnalystOpinions",
        "overallRisk", "auditRisk", "boardRisk", "compensationRisk",
        "fullTimeEmployees", "website",
    ]

    seen = set()
    for field in priority_fields:
        if field in info_dict:
            value = info_dict[field]
            # Only include simple values (not lists or nested dicts)
            if isinstance(value, (int, float, str, bool, type(None))):
                row[field] = value
                seen.add(field)

    # Then add any remaining fields we haven't included yet
    for field in sorted(info_dict.keys()):
        if field not in seen:
            value = info_dict[field]
            if isinstance(value, (int, float, str, bool, type(None))):
                row[field] = value

    return row


def style_excel_sheet(worksheet, max_column_width=55):
    """
    Apply formatting to an Excel sheet:
    - Navy header row with white bold text
    - Freeze the top row so headers stay visible when scrolling
    - Auto-fit column widths
    """
    # Header row style
    header_fill = PatternFill("solid", fgColor="1A2C4E")  # navy blue
    header_font = Font(bold=True, color="FFFFFF", size=9)
    for cell in worksheet[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(wrap_text=False, horizontal="center")

    # Body text style
    body_font = Font(size=9)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    # Freeze the header row
    worksheet.freeze_panes = "B2"

    # Auto-fit column widths
    for col in worksheet.columns:
        letter  = get_column_letter(col[0].column)
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=8,
        )
        worksheet.column_dimensions[letter].width = min(max_len + 2, max_column_width)


# ---------------------------------------------------------------------------
# DATA COLLECTION FUNCTIONS
# These do the actual work of fetching data from Yahoo Finance.
# You don't need to change anything in this section.
# ---------------------------------------------------------------------------

def get_daily_data(tickers, start_str, end_str):
    """
    Fetch daily price data (OHLCV + adjusted close + dividends + splits)
    for every ticker in the period.

    Returns three DataFrames:
      - daily_df   : one row per trading day per ticker
      - divs_df    : only the days a dividend was paid
      - splits_df  : only the days a split occurred
    """
    all_daily  = []
    all_divs   = []
    all_splits = []

    for ticker in tickers:
        print(f"  {ticker}...", end=" ", flush=True)
        try:
            raw = yf.Ticker(ticker).history(
                start=start_str,
                end=end_str,
                interval="1d",
                auto_adjust=False,  # keeps raw Close + separate Adj Close column
                actions=True,       # adds Dividends and Stock Splits columns
            )

            if raw is None or raw.empty:
                print("no data")
                continue

            # Clean up the index and column names
            raw.index      = remove_timezone(raw.index)
            raw.index.name = "Date"
            raw            = raw.reset_index()
            raw["Date"]    = raw["Date"].dt.date   # date-only, no time component in Excel
            raw.insert(0, "Ticker", ticker)

            raw.rename(columns={
                "Adj Close":    "Adj_Close",
                "Dividends":    "Dividend",
                "Stock Splits": "Split_Ratio",
            }, inplace=True)

            # Make sure dividend and split columns exist even if Yahoo omitted them
            for col in ("Dividend", "Split_Ratio"):
                if col not in raw.columns:
                    raw[col] = 0.0

            # Keep columns in a clean, consistent order
            ordered_cols = ["Ticker","Date","Open","High","Low","Close",
                            "Adj_Close","Volume","Dividend","Split_Ratio"]
            raw = raw[[c for c in ordered_cols if c in raw.columns]]

            all_daily.append(raw)

            # Pull out just the rows where a dividend was paid
            div_rows = raw[raw["Dividend"] > 0][
                ["Ticker","Date","Dividend","Close","Adj_Close"]
            ].copy()
            div_rows.rename(columns={"Dividend": "Dividend_Per_Share"}, inplace=True)
            if not div_rows.empty:
                all_divs.append(div_rows)

            # Pull out just the rows where a split occurred
            split_rows = raw[raw["Split_Ratio"] != 0][
                ["Ticker","Date","Split_Ratio","Close"]
            ].copy()
            if not split_rows.empty:
                all_splits.append(split_rows)

            # Print a brief summary for this ticker
            n      = len(raw)
            n_div  = (raw["Dividend"] > 0).sum()
            n_spl  = (raw["Split_Ratio"] != 0).sum()
            print(f"{n} days | {n_div} dividend(s) | {n_spl} split(s)")

        except Exception as e:
            print(f"ERROR — {e}")

    # Combine all tickers into single DataFrames
    def combine(lst):
        return pd.concat(lst, ignore_index=True) if lst else pd.DataFrame()

    return combine(all_daily), combine(all_divs), combine(all_splits)


def get_info_data(tickers):
    """
    Fetch the snapshot .info dictionary from Yahoo Finance for every ticker.
    Returns a DataFrame with one row per ticker and ~180 columns.
    """
    rows = []
    for ticker in tickers:
        print(f"  {ticker}...", end=" ", flush=True)
        try:
            info = yf.Ticker(ticker).info or {}
            rows.append(build_info_row(ticker, info))
            print(f"{len(info)} fields")
        except Exception as e:
            rows.append({"Ticker": ticker})
            print(f"ERROR — {e}")
    return pd.DataFrame(rows)


def get_quarterly_financials(tickers):
    """
    Fetch the last 4 quarters of income statements, balance sheets,
    and cash flow statements for every ticker.

    Returns three DataFrames (one per statement type), with all tickers stacked
    so each row = one quarter for one ticker.
    """
    inc_rows = []
    bal_rows = []
    cf_rows  = []

    # Map: yfinance attribute name → which list to append to → label for printing
    statements = [
        ("quarterly_income_stmt",   inc_rows, "Income"),
        ("quarterly_balance_sheet", bal_rows, "Balance"),
        ("quarterly_cashflow",       cf_rows, "CashFlow"),
    ]

    for ticker in tickers:
        print(f"  {ticker}...", end=" ", flush=True)
        t      = yf.Ticker(ticker)
        labels = []

        for attr_name, target_list, label in statements:
            try:
                df = getattr(t, attr_name, None)
                if df is None or df.empty:
                    continue

                # yfinance gives us: rows = metric names, columns = quarter dates
                # We transpose so each row = one quarter
                df_transposed = df.T.copy()
                df_transposed.index      = remove_timezone(df_transposed.index)
                df_transposed.index.name = "Quarter_End"
                df_transposed            = df_transposed.reset_index()
                df_transposed["Quarter_End"] = df_transposed["Quarter_End"].dt.date
                df_transposed.insert(0, "Ticker", ticker)

                # Convert all financial values to numbers
                for col in df_transposed.columns:
                    if col not in ("Ticker", "Quarter_End"):
                        df_transposed[col] = pd.to_numeric(df_transposed[col], errors="coerce")

                target_list.append(df_transposed)
                labels.append(f"{label}({len(df_transposed)}Q)")

            except Exception as e:
                labels.append(f"{label}(err)")

        print(", ".join(labels) if labels else "no financials found")

    def combine(lst):
        return pd.concat(lst, ignore_index=True) if lst else pd.DataFrame()

    return combine(inc_rows), combine(bal_rows), combine(cf_rows)


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# Writes all the data to a single Excel file with multiple sheets.
# You don't need to change anything in this section.
# ---------------------------------------------------------------------------

# Empty placeholder DataFrames used when no events occurred in the period
EMPTY_DIVIDENDS = pd.DataFrame(
    columns=["Ticker","Date","Dividend_Per_Share","Close","Adj_Close"]
)
EMPTY_SPLITS = pd.DataFrame(
    columns=["Ticker","Date","Split_Ratio","Close"]
)


def save_to_excel(path, daily, divs, splits, info, income, balance, cashflow):
    """Write all DataFrames to a formatted Excel file."""

    # Use placeholder sheets if there were no dividends or splits in the period
    div_sheet    = divs   if (divs   is not None and not divs.empty)   else EMPTY_DIVIDENDS
    splits_sheet = splits if (splits is not None and not splits.empty) else EMPTY_SPLITS

    sheets = [
        ("Info & Ratios",     info),
        ("Daily Prices",      daily),
        ("Dividends",         div_sheet),
        ("Stock Splits",      splits_sheet),
        ("Income Stmt (Q)",   income),
        ("Balance Sheet (Q)", balance),
        ("Cash Flow (Q)",     cashflow),
    ]

    # Write all sheets to Excel
    with pd.ExcelWriter(
        path,
        engine="openpyxl",
        date_format="YYYY-MM-DD",
        datetime_format="YYYY-MM-DD",
    ) as writer:
        for sheet_name, df in sheets:
            if df is None or not isinstance(df, pd.DataFrame):
                continue
            # Strip timezone from dates before writing
            df = remove_timezone_from_df(df)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Re-open the file to apply formatting (openpyxl doesn't allow this during write)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        style_excel_sheet(ws)
    wb.save(path)


# ---------------------------------------------------------------------------
# MAIN — this is what runs when you execute the script
# ---------------------------------------------------------------------------

def main():
    # Allow optional command-line overrides:
    #   python collect_data.py --start 2025-01-01 --end 2025-12-31
    #   python collect_data.py --out my_custom_filename.xlsx
    parser = argparse.ArgumentParser(
        description="Download Yahoo Finance data for the JDCA portfolio"
    )
    parser.add_argument("--start", default=None, help="Start date YYYY-MM-DD")
    parser.add_argument("--end",   default=None, help="End date   YYYY-MM-DD")
    parser.add_argument("--out",   default=None, help="Output filename (.xlsx)")
    args = parser.parse_args()

    # Use command-line dates if provided, otherwise fall back to the defaults at the top
    start = datetime.strptime(args.start, "%Y-%m-%d") if args.start else START_DATE
    end   = datetime.strptime(args.end,   "%Y-%m-%d") if args.end   else END_DATE

    # yfinance's end date is exclusive (it stops before it), so we add one day
    start_str = start.strftime("%Y-%m-%d")
    end_str   = (end + timedelta(days=1)).strftime("%Y-%m-%d")

    outfile = args.out or OUTPUT_FILE

    # Print a summary of what's about to run
    print("=" * 60)
    print("JDCA Portfolio Data Collector")
    print(f"  Period  : {start_str}  to  {end.strftime('%Y-%m-%d')}")
    print(f"  Tickers : {len(TICKERS)}")
    print(f"  Output  : {outfile}")
    print("=" * 60)

    # Step 1: Daily prices + dividends + splits
    print("\n[1/3]  Daily prices, dividends & splits")
    daily, divs, splits = get_daily_data(TICKERS, start_str, end_str)

    # Step 2: Snapshot info fields
    print("\n[2/3]  Snapshot info & ratios")
    info = get_info_data(TICKERS)

    # Step 3: Quarterly financials
    print("\n[3/3]  Quarterly financials (most recent filings)")
    income, balance, cashflow = get_quarterly_financials(TICKERS)

    # Save everything to Excel
    print(f"\nSaving to {outfile} ...")
    save_to_excel(outfile, daily, divs, splits, info, income, balance, cashflow)

    # Final summary
    print("\nAll done!")
    print(f"  Daily rows      : {len(daily):>6,}")
    print(f"  Dividend events : {len(divs):>6,}  (in the period)")
    print(f"  Split events    : {len(splits):>6,}  (in the period)")
    print(f"  Info columns    : {info.shape[1]:>6,}")
    print(f"  Income rows     : {len(income):>6,}")
    print(f"  Balance rows    : {len(balance):>6,}")
    print(f"  CashFlow rows   : {len(cashflow):>6,}")
    print(f"\nFile saved: {outfile}")


# This line means: only run main() if this file is executed directly
# (not if it's imported by another script)
if __name__ == "__main__":
    main()
